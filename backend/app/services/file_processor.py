"""
📄 FILE PROCESSOR + PILI v3.0 - OCR INTELIGENTE MULTIMODAL
📁 RUTA: backend/app/services/file_processor.py

PILI (Procesadora Inteligente de Licitaciones Industriales) integrada con 
procesamiento de archivos para OCR inteligente y extracción especializada.

🎯 NUEVAS CARACTERÍSTICAS PILI v3.0:
- OCR inteligente especializado por tipo de documento
- Procesamiento multimodal (fotos, PDFs, Word, Excel, manuscritos)
- Extracción selectiva de información relevante por servicio
- Análisis automático de contenido técnico eléctrico
- Detección de elementos específicos (planos, especificaciones, etc.)
- Integración con agentes PILI especializados

🔄 CONSERVA TODO LO EXISTENTE:
- procesar_archivo() ✅
- Todos los extractores por tipo ✅  
- validar_archivo() ✅
- guardar_archivo() ✅
- extraer_metadata() ✅
- Todo el manejo de errores robusto ✅

🔧 REPARADO:
- Función _detectar_tipo_mime faltante agregada
- Cambio de python-magic a filetype para compatibilidad Windows
- Manejo robusto de errores en detección de tipos MIME
"""

import os
import shutil
from typing import Dict, Any, Optional, List
from datetime import datetime
import logging
import json

# Imports para procesamiento de archivos (conservados)
import PyPDF2
from docx import Document
from openpyxl import load_workbook
from PIL import Image
import pytesseract
import filetype  # Reemplazado magic por filetype para compatibilidad Windows

# Imports PILI
from pathlib import Path
import re

# Configuración (conservada)
from app.core.config import settings

logger = logging.getLogger(__name__)

class FileProcessor:
    """
    🔄 PROCESADOR ORIGINAL CONSERVADO + 🤖 PILI INTEGRADA
    
    Mantiene toda la funcionalidad existente pero agrega capacidades
    inteligentes de PILI para procesamiento especializado por servicio.
    """
    
    def __init__(self):
        """🔄 CONSERVADO + 🤖 PILI mejorado"""
        
        # Configuración original conservada
        self.upload_dir = settings.UPLOAD_DIR
        self.allowed_extensions = settings.ALLOWED_EXTENSIONS
        self.max_file_size = settings.MAX_UPLOAD_SIZE_MB
        
        # 🤖 Configuración PILI especializada
        self.pili_patterns = {
            "electricas": [
                r"instalaci[oó]n\s+el[eé]ctrica",
                r"punto\s+de\s+luz",
                r"tomacorriente",
                r"tablero\s+el[eé]ctrico",
                r"cable\s+THW",
                r"pozo\s+a\s+tierra",
                r"CNE",
                r"voltaje",
                r"amperaje",
                r"circuito",
                r"interruptor",
                r"carga\s+el[eé]ctrica"
            ],
            "automatizacion": [
                r"automatizaci[oó]n",
                r"control\s+automatico",
                r"sensor",
                r"PLC",
                r"SCADA",
                r"domótica",
                r"motor",
                r"actuador"
            ],
            "cctv": [
                r"c[aá]mara",
                r"CCTV",
                r"vigilancia",
                r"monitoreo",
                r"grabaci[oó]n",
                r"DVR",
                r"IP\s+camera"
            ],
            "redes": [
                r"red\s+inform[aá]tica",
                r"ethernet",
                r"switch",
                r"router",
                r"fibra\s+[oó]ptica",
                r"cableado\s+estructurado",
                r"rack"
            ],
            "normativas": [
                r"CNE",
                r"CNE-Utilizaci[oó]n",
                r"CNE-Suministro", 
                r"RNE",
                r"INDECOPI",
                r"OSINERGMIN",
                r"DS-066-2007",
                r"IEEE-80",
                r"NFPA"
            ]
        }
        
        logger.info("🤖 FileProcessor + PILI v3.0 inicializado con OCR inteligente")

    # ════════════════════════════════════════════════════════════════
    # 🤖 PILI - PROCESAMIENTO INTELIGENTE POR SERVICIO
    # ════════════════════════════════════════════════════════════════

    def procesar_con_pili(
        self,
        archivo_path: str,
        nombre_original: str,
        tipo_servicio: str = "electricidad",
        agente_pili: str = "PILI Analista"
    ) -> Dict[str, Any]:
        """
        🤖 NUEVO PILI v3.0 - Procesamiento inteligente especializado
        
        Procesa archivos con análisis específico según el tipo de servicio,
        usando patrones especializados y extracción selectiva de información.
        
        Args:
            archivo_path: Ruta del archivo a procesar
            nombre_original: Nombre original del archivo
            tipo_servicio: Tipo de servicio (electricidad, automatizacion, cctv, redes)
            agente_pili: Agente PILI responsable del procesamiento
            
        Returns:
            Dict con análisis inteligente y datos estructurados
        """
        
        try:
            logger.info(f"🤖 {agente_pili} procesando {nombre_original} para servicio: {tipo_servicio}")
            
            # Procesamiento base usando método original
            resultado_base = self.procesar_archivo(archivo_path, nombre_original)
            
            if not resultado_base["exito"]:
                return {
                    "exito": False,
                    "error": f"Error en procesamiento base: {resultado_base.get('error', 'Desconocido')}",
                    "agente_pili": agente_pili
                }
            
            # Análisis especializado PILI
            contenido_texto = resultado_base.get("contenido_texto", "")
            
            analisis_pili = {
                "agente_responsable": agente_pili,
                "tipo_servicio": tipo_servicio,
                "timestamp": datetime.now().isoformat(),
                
                # Análisis de contenido técnico
                "elementos_detectados": self._detectar_elementos_tecnicos(contenido_texto, tipo_servicio),
                "normativas_identificadas": self._identificar_normativas(contenido_texto),
                "datos_estructurados": self._extraer_datos_estructurados(contenido_texto, tipo_servicio),
                
                # Clasificación inteligente
                "tipo_documento_identificado": self._clasificar_documento_tecnico(contenido_texto, nombre_original),
                "nivel_complejidad": self._evaluar_complejidad(contenido_texto),
                "recomendaciones_procesamiento": self._generar_recomendaciones(contenido_texto, tipo_servicio),
                
                # Extracción específica por tipo de servicio
                "datos_especificos": self._extraer_por_servicio(contenido_texto, tipo_servicio),
                
                # Metadatos enriquecidos
                "confianza_extraccion": self._calcular_confianza(contenido_texto),
                "requiere_revision_humana": self._necesita_revision(contenido_texto),
                "siguiente_paso_recomendado": self._recomendar_siguiente_paso(contenido_texto, tipo_servicio)
            }
            
            # Combinar resultado base con análisis PILI
            resultado_final = {
                **resultado_base,
                "pili_analysis": analisis_pili,
                "procesamiento_inteligente": True,
                "mensaje_pili": f"✅ {agente_pili} completó el análisis del documento. Elementos técnicos detectados: {len(analisis_pili['elementos_detectados'])}"
            }
            
            logger.info(f"✅ {agente_pili} completó procesamiento inteligente de {nombre_original}")
            return resultado_final
            
        except Exception as e:
            logger.error(f"❌ Error en procesamiento PILI: {str(e)}")
            return {
                "exito": False,
                "error": f"Error PILI: {str(e)}",
                "agente_pili": agente_pili,
                "procesamiento_inteligente": False
            }

    def _detectar_elementos_tecnicos(self, contenido: str, tipo_servicio: str) -> List[Dict[str, Any]]:
        """Detecta elementos técnicos específicos según el servicio"""
        
        elementos = []
        patrones = self.pili_patterns.get(tipo_servicio, [])
        
        for patron in patrones:
            matches = re.finditer(patron, contenido, re.IGNORECASE | re.UNICODE)
            for match in matches:
                elementos.append({
                    "elemento": match.group(),
                    "posicion": match.start(),
                    "contexto": contenido[max(0, match.start()-50):match.end()+50],
                    "patron_usado": patron,
                    "tipo_servicio": tipo_servicio
                })
        
        return elementos

    def _identificar_normativas(self, contenido: str) -> List[str]:
        """Identifica normativas técnicas mencionadas en el documento"""
        
        normativas_encontradas = []
        patrones_normativas = self.pili_patterns["normativas"]
        
        for patron in patrones_normativas:
            if re.search(patron, contenido, re.IGNORECASE):
                # Extraer la normativa específica
                matches = re.findall(patron, contenido, re.IGNORECASE)
                normativas_encontradas.extend(matches)
        
        # Eliminar duplicados y ordenar
        return sorted(list(set(normativas_encontradas)))

    def _extraer_datos_estructurados(self, contenido: str, tipo_servicio: str) -> Dict[str, Any]:
        """Extrae datos estructurados específicos del documento"""
        
        datos = {
            "cantidades": self._extraer_cantidades(contenido),
            "medidas": self._extraer_medidas(contenido),
            "precios": self._extraer_precios(contenido),
            "especificaciones": self._extraer_especificaciones_tecnicas(contenido, tipo_servicio)
        }
        
        return datos

    def _extraer_cantidades(self, contenido: str) -> List[Dict[str, Any]]:
        """Extrae cantidades numéricas del documento"""
        
        # Patrones para cantidades
        patrones_cantidad = [
            r"(\d+)\s*(punto[s]?\s+de\s+luz)",
            r"(\d+)\s*(tomacorriente[s]?)",
            r"(\d+)\s*(metro[s]?\s+de\s+cable)",
            r"(\d+)\s*(m[2]?)",
            r"(\d+)\s*(und|unidad[es]?)",
            r"(\d+)\s*(pza[s]?|pieza[s]?)"
        ]
        
        cantidades = []
        for patron in patrones_cantidad:
            matches = re.finditer(patron, contenido, re.IGNORECASE)
            for match in matches:
                cantidades.append({
                    "cantidad": int(match.group(1)),
                    "item": match.group(2),
                    "contexto": match.group(0)
                })
        
        return cantidades

    def _extraer_medidas(self, contenido: str) -> List[Dict[str, Any]]:
        """Extrae medidas y dimensiones del documento"""
        
        # Patrones para medidas
        patrones_medidas = [
            r"(\d+\.?\d*)\s*x\s*(\d+\.?\d*)\s*(m|metro[s]?|cm|centímetro[s]?)",
            r"(\d+\.?\d*)\s*(m[2]|metro[s]?\s+cuadrado[s]?)",
            r"(\d+\.?\d*)\s*(mm[2]|AWG|THW)",
            r"(\d+\.?\d*)\s*(voltio[s]?|V|amperio[s]?|A|watt[s]?|W)"
        ]
        
        medidas = []
        for patron in patrones_medidas:
            matches = re.finditer(patron, contenido, re.IGNORECASE)
            for match in matches:
                medidas.append({
                    "valor": match.group(1),
                    "unidad": match.group(2) if len(match.groups()) >= 2 else "",
                    "contexto": match.group(0)
                })
        
        return medidas

    def _extraer_precios(self, contenido: str) -> List[Dict[str, Any]]:
        """Extrae precios y costos del documento"""
        
        # Patrones para precios en soles peruanos
        patrones_precios = [
            r"S/\.?\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)",
            r"(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)\s*soles",
            r"PEN\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)",
            r"(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)\s*nuevos soles"
        ]
        
        precios = []
        for patron in patrones_precios:
            matches = re.finditer(patron, contenido, re.IGNORECASE)
            for match in matches:
                valor_str = match.group(1).replace(",", "")
                try:
                    valor = float(valor_str)
                    precios.append({
                        "valor": valor,
                        "moneda": "PEN",
                        "formato_original": match.group(0),
                        "contexto": contenido[max(0, match.start()-30):match.end()+30]
                    })
                except ValueError:
                    pass
        
        return precios

    def _extraer_especificaciones_tecnicas(self, contenido: str, tipo_servicio: str) -> Dict[str, List[str]]:
        """Extrae especificaciones técnicas específicas del servicio"""
        
        especificaciones = {
            "cables": [],
            "equipos": [],
            "protecciones": [],
            "mediciones": [],
            "normas_aplicables": []
        }
        
        if tipo_servicio == "electricidad":
            # Especificaciones eléctricas
            especificaciones["cables"] = re.findall(r"cable\s+THW\s+\d+\.?\d*\s*mm[2]?", contenido, re.IGNORECASE)
            especificaciones["equipos"] = re.findall(r"tablero\s+\w+", contenido, re.IGNORECASE)
            especificaciones["protecciones"] = re.findall(r"interruptor\s+\d+A?", contenido, re.IGNORECASE)
            
        elif tipo_servicio == "automatizacion":
            # Especificaciones de automatización
            especificaciones["equipos"] = re.findall(r"PLC\s+\w+", contenido, re.IGNORECASE)
            especificaciones["equipos"].extend(re.findall(r"sensor\s+\w+", contenido, re.IGNORECASE))
            
        elif tipo_servicio == "cctv":
            # Especificaciones de CCTV
            especificaciones["equipos"] = re.findall(r"cámara\s+\w+", contenido, re.IGNORECASE)
            especificaciones["equipos"].extend(re.findall(r"DVR\s+\w+", contenido, re.IGNORECASE))
            
        return especificaciones

    def _clasificar_documento_tecnico(self, contenido: str, nombre_archivo: str) -> str:
        """Clasifica el tipo de documento técnico"""
        
        # Análisis por nombre de archivo
        nombre_lower = nombre_archivo.lower()
        
        if any(word in nombre_lower for word in ["plano", "drawing", "dwg"]):
            return "Plano técnico"
        elif any(word in nombre_lower for word in ["especificacion", "spec", "manual"]):
            return "Especificación técnica"
        elif any(word in nombre_lower for word in ["cotizacion", "quote", "presupuesto"]):
            return "Cotización/Presupuesto"
        elif any(word in nombre_lower for word in ["memoria", "calculo", "calculation"]):
            return "Memoria de cálculo"
        elif any(word in nombre_lower for word in ["informe", "report"]):
            return "Informe técnico"
        
        # Análisis por contenido
        if re.search(r"memoria\s+de\s+c[aá]lculo", contenido, re.IGNORECASE):
            return "Memoria de cálculo"
        elif re.search(r"especificaci[oó]n\s+t[eé]cnica", contenido, re.IGNORECASE):
            return "Especificación técnica"
        elif re.search(r"presupuesto|cotizaci[oó]n", contenido, re.IGNORECASE):
            return "Cotización/Presupuesto"
        
        return "Documento técnico general"

    def _evaluar_complejidad(self, contenido: str) -> str:
        """Evalúa el nivel de complejidad del documento"""
        
        # Factores de complejidad
        palabras_tecnicas = len(re.findall(r"\b(?:instalaci[oó]n|circuito|tablero|cable|interruptor|voltaje)\b", contenido, re.IGNORECASE))
        numeros_especificos = len(re.findall(r"\d+\.?\d*\s*(?:mm|A|V|W|m)", contenido, re.IGNORECASE))
        referencias_normativas = len(re.findall(r"\b(?:CNE|IEEE|NFPA|DS-066)\b", contenido, re.IGNORECASE))
        
        # Calcular puntuación de complejidad
        puntuacion = palabras_tecnicas * 2 + numeros_especificos + referencias_normativas * 3
        
        if puntuacion >= 20:
            return "Alta complejidad"
        elif puntuacion >= 10:
            return "Complejidad media"
        else:
            return "Baja complejidad"

    def _generar_recomendaciones(self, contenido: str, tipo_servicio: str) -> List[str]:
        """Genera recomendaciones de procesamiento"""
        
        recomendaciones = []
        
        # Recomendaciones basadas en contenido
        if re.search(r"instalaci[oó]n\s+el[eé]ctrica", contenido, re.IGNORECASE):
            recomendaciones.append("Revisar especificaciones de instalación eléctrica")
            
        if re.search(r"pozo\s+a\s+tierra", contenido, re.IGNORECASE):
            recomendaciones.append("Verificar normativa de puesta a tierra")
            
        if re.search(r"tablero\s+el[eé]ctrico", contenido, re.IGNORECASE):
            recomendaciones.append("Validar dimensionamiento del tablero eléctrico")
            
        # Recomendaciones por tipo de servicio
        if tipo_servicio == "electricidad":
            recomendaciones.append("Aplicar normativa CNE-Utilización vigente")
        elif tipo_servicio == "automatizacion":
            recomendaciones.append("Considerar protocolos de comunicación industrial")
        elif tipo_servicio == "cctv":
            recomendaciones.append("Evaluar calidad de imagen y almacenamiento")
        
        return recomendaciones

    def _extraer_por_servicio(self, contenido: str, tipo_servicio: str) -> Dict[str, Any]:
        """Extrae datos específicos según el tipo de servicio"""
        
        if tipo_servicio == "electricidad":
            return {
                "puntos_luz": len(re.findall(r"punto\s+de\s+luz", contenido, re.IGNORECASE)),
                "tomacorrientes": len(re.findall(r"tomacorriente", contenido, re.IGNORECASE)),
                "tableros": re.findall(r"tablero\s+[^,\n]*", contenido, re.IGNORECASE),
                "cables_detectados": re.findall(r"cable\s+THW[^,\n]*", contenido, re.IGNORECASE)
            }
        elif tipo_servicio == "automatizacion":
            return {
                "plcs": re.findall(r"PLC[^,\n]*", contenido, re.IGNORECASE),
                "sensores": re.findall(r"sensor[^,\n]*", contenido, re.IGNORECASE),
                "actuadores": re.findall(r"actuador[^,\n]*", contenido, re.IGNORECASE)
            }
        elif tipo_servicio == "cctv":
            return {
                "camaras": re.findall(r"c[aá]mara[^,\n]*", contenido, re.IGNORECASE),
                "dvrs": re.findall(r"DVR[^,\n]*", contenido, re.IGNORECASE),
                "monitores": re.findall(r"monitor[^,\n]*", contenido, re.IGNORECASE)
            }
        
        return {}

    def _calcular_confianza(self, contenido: str) -> float:
        """Calcula el nivel de confianza en la extracción de datos"""
        
        # Factores que aumentan la confianza
        palabras_tecnicas = len(re.findall(r"\b(?:instalaci[oó]n|circuito|cable|voltaje|amperaje)\b", contenido, re.IGNORECASE))
        numeros_con_unidades = len(re.findall(r"\d+\s*(?:mm|A|V|W|m)", contenido, re.IGNORECASE))
        estructura_documento = 1 if re.search(r"(?:especificaci[oó]n|descripci[oó]n|alcance)", contenido, re.IGNORECASE) else 0
        
        # Calcular confianza (0-100%)
        puntuacion = min(100, (palabras_tecnicas * 5 + numeros_con_unidades * 3 + estructura_documento * 10))
        
        return puntuacion / 100

    def _necesita_revision(self, contenido: str) -> bool:
        """Determina si el documento necesita revisión humana"""
        
        # Criterios que requieren revisión humana
        texto_muy_corto = len(contenido.strip()) < 100
        sin_datos_tecnicos = len(re.findall(r"\d+", contenido)) < 3
        texto_ilegible = contenido.count("?") > len(contenido) * 0.05
        
        return texto_muy_corto or sin_datos_tecnicos or texto_ilegible

    def _recomendar_siguiente_paso(self, contenido: str, tipo_servicio: str) -> str:
        """Recomienda el siguiente paso en el procesamiento"""
        
        if self._necesita_revision(contenido):
            return "Revisión humana requerida - documento con poca información extraíble"
        
        if re.search(r"cotizaci[oó]n|presupuesto", contenido, re.IGNORECASE):
            return "Generar cotización formal usando los datos extraídos"
        
        if re.search(r"proyecto|instalaci[oó]n", contenido, re.IGNORECASE):
            return "Crear estructura de proyecto y cronograma de trabajo"
        
        if re.search(r"especificaci[oó]n|memoria", contenido, re.IGNORECASE):
            return "Generar informe técnico basado en las especificaciones"
        
        return "Continuar con el flujo de trabajo estándar"

    # ════════════════════════════════════════════════════════════════
    # 🔄 MÉTODOS ORIGINALES CONSERVADOS
    # ════════════════════════════════════════════════════════════════

    def procesar_archivo(self, archivo_path: str, nombre_original: str) -> Dict[str, Any]:
        """
        🔄 CONSERVADO - Procesa un archivo y extrae su contenido
        
        Método original mantenido para compatibilidad hacia atrás.
        """
        
        # Detectar tipo de archivo
        tipo_mime = self._detectar_tipo_mime(archivo_path)
        extension = nombre_original.split('.')[-1].lower()
        
        resultado = {
            "exito": False,
            "nombre_original": nombre_original,
            "tipo_mime": tipo_mime,
            "extension": extension,
            "contenido_texto": "",
            "metadata": {},
            "error": None
        }
        
        try:
            # Procesar según tipo
            if extension == 'pdf':
                contenido = self._extraer_texto_pdf(archivo_path)
            elif extension in ['docx', 'doc']:
                contenido = self._extraer_texto_word(archivo_path)
            elif extension in ['xlsx', 'xls']:
                contenido = self._extraer_texto_excel(archivo_path)
            elif extension in ['jpg', 'jpeg', 'png', 'bmp']:
                contenido = self._extraer_texto_imagen(archivo_path)
            elif extension == 'txt':
                contenido = self._extraer_texto_plano(archivo_path)
            else:
                contenido = f"Tipo de archivo no soportado para extracción: {extension}"
            
            resultado.update({
                "exito": True,
                "contenido_texto": contenido,
                "metadata": self.extraer_metadata(archivo_path),
                "tipo_documento_detectado": self._detectar_tipo_documento(contenido, nombre_original),
            })
            
        except Exception as e:
            logger.error(f"Error procesando archivo {nombre_original}: {str(e)}")
            resultado["error"] = str(e)
        
        return resultado

    def _detectar_tipo_mime(self, archivo_path: str) -> str:
        """
        🔧 FUNCIÓN REPARADA - Detecta tipo MIME usando filetype (compatible Windows)
        
        Reemplaza el uso de python-magic por filetype para mejor compatibilidad.
        """
        try:
            # Usar filetype para detectar tipo MIME (compatible con Windows)
            kind = filetype.guess(archivo_path)
            if kind is not None:
                return kind.mime
            
            # Fallback: detectar por extensión
            extension = archivo_path.split('.')[-1].lower()
            mime_types = {
                'pdf': 'application/pdf',
                'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                'doc': 'application/msword',
                'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'xls': 'application/vnd.ms-excel',
                'txt': 'text/plain',
                'jpg': 'image/jpeg',
                'jpeg': 'image/jpeg',
                'png': 'image/png',
                'bmp': 'image/bmp',
                'gif': 'image/gif'
            }
            return mime_types.get(extension, 'application/octet-stream')
            
        except Exception as e:
            logger.warning(f"Error detectando tipo MIME para {archivo_path}: {e}")
            # Fallback seguro por extensión
            extension = archivo_path.split('.')[-1].lower()
            return {
                'pdf': 'application/pdf',
                'docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                'txt': 'text/plain'
            }.get(extension, 'application/octet-stream')

    def _detectar_tipo_documento(self, contenido: str, nombre_archivo: str) -> Dict[str, Any]:
        """
        🔄 CONSERVADO - Detecta el tipo de documento basado en contenido
        """
        tipos_detectados = []
        
        # Análisis por nombre de archivo
        nombre_lower = nombre_archivo.lower()
        if any(word in nombre_lower for word in ["cotizacion", "presupuesto", "quote"]):
            tipos_detectados.append("cotizacion")
        if any(word in nombre_lower for word in ["proyecto", "project"]):
            tipos_detectados.append("proyecto")
        if any(word in nombre_lower for word in ["informe", "reporte", "report"]):
            tipos_detectados.append("informe")
        if any(word in nombre_lower for word in ["plano", "drawing"]):
            tipos_detectados.append("plano")
        if any(word in nombre_lower for word in ["especificacion", "spec"]):
            tipos_detectados.append("especificacion")
        
        # Análisis por contenido
        if re.search(r"cotizaci[oó]n|presupuesto", contenido, re.IGNORECASE):
            tipos_detectados.append("cotizacion")
        if re.search(r"proyecto|instalaci[oó]n", contenido, re.IGNORECASE):
            tipos_detectados.append("proyecto")
        if re.search(r"informe|reporte|conclusi[oó]n", contenido, re.IGNORECASE):
            tipos_detectados.append("informe")
        if re.search(r"especificaci[oó]n\s+t[eé]cnica", contenido, re.IGNORECASE):
            tipos_detectados.append("especificacion")
        
        # Detectar si es documento técnico eléctrico
        es_tecnico_electrico = any([
            re.search(r"instalaci[oó]n\s+el[eé]ctrica", contenido, re.IGNORECASE),
            re.search(r"punto\s+de\s+luz", contenido, re.IGNORECASE),
            re.search(r"tablero\s+el[eé]ctrico", contenido, re.IGNORECASE),
            re.search(r"cable\s+THW", contenido, re.IGNORECASE),
            re.search(r"CNE", contenido, re.IGNORECASE)
        ])
        
        return {
            "tipos_detectados": list(set(tipos_detectados)),
            "tipo_principal": tipos_detectados[0] if tipos_detectados else "documento_general",
            "es_tecnico_electrico": es_tecnico_electrico,
            "confianza": min(1.0, len(tipos_detectados) / 2)
        }

    def _extraer_texto_pdf(self, archivo_path: str) -> str:
        """
        🔄 CONSERVADO - Extrae texto de un archivo PDF
        """
        texto = []
        
        with open(archivo_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                texto.append(page.extract_text())
        
        return '\n'.join(texto)
    
    def _extraer_texto_word(self, archivo_path: str) -> str:
        """
        🔄 CONSERVADO - Extrae texto de un archivo Word
        """
        documento = Document(archivo_path)
        return '\n'.join([p.text for p in documento.paragraphs])
    
    def _extraer_texto_excel(self, archivo_path: str) -> str:
        """
        🔄 CONSERVADO - Extrae texto de un archivo Excel
        """
        workbook = load_workbook(archivo_path)
        texto = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_text = ' '.join([str(cell) for cell in row if cell is not None])
                if row_text.strip():
                    texto.append(row_text)
        
        return '\n'.join(texto)
    
    def _extraer_texto_imagen(self, archivo_path: str) -> str:
        """
        🔄 CONSERVADO - Extrae texto de una imagen usando OCR
        """
        try:
            imagen = Image.open(archivo_path)
            texto = pytesseract.image_to_string(imagen, lang='spa+eng')
            return texto
        except Exception as e:
            logger.warning(f"Error en OCR para {archivo_path}: {e}")
            return f"Error extracting text from image: {str(e)}"
    
    def _extraer_texto_plano(self, archivo_path: str) -> str:
        """
        🔄 CONSERVADO - Extrae texto de un archivo de texto plano
        """
        with open(archivo_path, 'r', encoding='utf-8') as file:
            return file.read()

    def validar_archivo(self, archivo) -> Dict[str, Any]:
        """
        🔄 CONSERVADO - Valida un archivo antes del procesamiento
        """
        resultado = {
            "valido": False,
            "error": None,
            "tamaño_mb": 0,
            "extension": "",
            "tipo_mime": ""
        }
        
        try:
            # Validar extensión
            if hasattr(archivo, 'filename'):
                nombre = archivo.filename
            else:
                nombre = str(archivo)
                
            extension = nombre.split('.')[-1].lower()
            
            if extension not in self.allowed_extensions:
                resultado["error"] = f"Extensión {extension} no permitida"
                return resultado
            
            # Validar tamaño
            if hasattr(archivo, 'size'):
                tamaño_mb = archivo.size / (1024 * 1024)
            else:
                tamaño_mb = os.path.getsize(archivo) / (1024 * 1024)
            
            if tamaño_mb > self.max_file_size:
                resultado["error"] = f"Archivo muy grande: {tamaño_mb:.2f}MB (máximo: {self.max_file_size}MB)"
                return resultado
            
            resultado.update({
                "valido": True,
                "tamaño_mb": round(tamaño_mb, 2),
                "extension": extension
            })
            
        except Exception as e:
            resultado["error"] = f"Error validando archivo: {str(e)}"
        
        return resultado

    def extraer_metadata(self, archivo_path: str) -> Dict[str, Any]:
        """
        🔄 CONSERVADO - Extrae metadatos de un archivo
        """
        try:
            stat_info = os.stat(archivo_path)
            
            return {
                "tamaño_bytes": stat_info.st_size,
                "tamaño_mb": round(stat_info.st_size / (1024 * 1024), 2),
                "fecha_modificacion": datetime.fromtimestamp(stat_info.st_mtime).isoformat(),
                "fecha_acceso": datetime.fromtimestamp(stat_info.st_atime).isoformat(),
                "extension": archivo_path.split('.')[-1].lower(),
                "ruta": archivo_path
            }
        except Exception as e:
            logger.error(f"Error extrayendo metadatos: {e}")
            return {}

    def guardar_archivo(self, archivo, nombre_unico: str) -> str:
        """
        🔄 CONSERVADO - Guarda un archivo en el sistema
        """
        ruta_completa = os.path.join(self.upload_dir, nombre_unico)
        
        with open(ruta_completa, "wb") as buffer:
            shutil.copyfileobj(archivo.file, buffer)
        
        return ruta_completa
    
    def eliminar_archivo(self, ruta: str) -> bool:
        """
        🔄 CONSERVADO - Elimina un archivo del sistema
        """
        try:
            if os.path.exists(ruta):
                os.remove(ruta)
                return True
            return False
        except Exception as e:
            logger.error(f"Error al eliminar archivo: {e}")
            return False

# ═══════════════════════════════════════════════════════════════
# 🎯 INSTANCIA GLOBAL MEJORADA CON PILI
# ═══════════════════════════════════════════════════════════════

# Crear instancia global con manejo robusto de errores
try:
    file_processor = FileProcessor()
    logger.info("✅ FileProcessor + PILI inicializado correctamente")
except Exception as e:
    logger.error(f"❌ Error crítico inicializando FileProcessor: {e}")
    file_processor = None

# Función auxiliar para obtener instancia segura
def get_file_processor():
    """Obtiene instancia de FileProcessor de forma segura"""
    global file_processor
    if file_processor is None:
        try:
            file_processor = FileProcessor()
        except Exception as e:
            logger.error(f"Error creando FileProcessor: {e}")
    return file_processor